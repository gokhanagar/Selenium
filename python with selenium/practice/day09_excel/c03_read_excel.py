'''
import openpyxl
# to load the workbook with its path
bk = openpyxl.load_workbook(“C:\\STM\\Python.xlsx”)
# to identify active worksheet
s = bk.active
# to identify maximum rows count
print ( s.max_row)
# to identify maximum columns count
print ( s.max_column)

===================================================================


import openpyxl
# to load the workbook with its path
bk = openpyxl.load_workbook(“C:\\STM\\Python.xlsx”)
# to identify active worksheet
s = bk.active
# to identify maximum rows count
print ( s.max_row)
# iterate till the count of occupied rows
for m in range ( 1, s.max_row + 1):
# iterate till the count of occupied columns
for n in range ( 1, s.max_column + 1):
# to get the cell data and print
print (s.cell (row=m, column=n). value) 


===================================================================


import openpyxl
# to load the workbook with its path
bk = openpyxl.load_workbook(“C:\\STM\\Python.xlsx”)
# to identify active worksheet
s = bk.active
# to identify maximum rows count
print ( s.max_row)
# iterate till the count of occupied rows
for m in range ( 1, s.max_row + 1): 
# to get all the cell values from column 1
if s.cell (row=m, column=1).value == "Python" :
# iterate till the count of occupied columns
for n in range ( 1, s.max_column + 1):
# to get all the cell data from a row and print
print (s.cell (row=m, column=n). value) 






















'''